# =============================================================================
# importar_base.py — Importa novas NFs do arquivo_base para o Quadro de Envios
# =============================================================================
# CHAVE: Data Emissão + NF + CNPJ Destinatário + Empresa
# Se a combinação NÃO existir no Quadro → adiciona no final
# Se JÁ existir → ignora (preserva campos manuais do operador)
# =============================================================================

import os, sys, shutil, logging
from datetime import datetime
import pandas as pd
import openpyxl

# ── CONFIGURAÇÃO ──
# Detecta automaticamente o OneDrive em qualquer máquina
def _encontrar_pasta():
    """Localiza a pasta 10-DASHIBOARDs no OneDrive de qualquer usuário."""
    import os
    # Tentar variáveis de ambiente do OneDrive (Windows)
    candidatos = [
        os.environ.get("OneDriveCommercial", ""),
        os.environ.get("OneDrive", ""),
        os.path.join(os.environ.get("USERPROFILE", ""), "OneDrive - Instituto Alfa e Beto"),
        os.path.join(os.environ.get("USERPROFILE", ""), "OneDrive"),
    ]
    sufixo = os.path.join("Logística-IAB", "10 - DASHIBOARDs")
    for base in candidatos:
        if not base:
            continue
        caminho = os.path.join(base, sufixo) if "Logística" not in base else os.path.join(base, "10 - DASHIBOARDs")
        if os.path.isdir(caminho):
            return caminho
        # Tentar direto
        caminho2 = os.path.join(base, "Logística-IAB", "10 - DASHIBOARDs")
        if os.path.isdir(caminho2):
            return caminho2
    # Fallback: pasta 2 níveis acima do script (etl/ -> dashboard_logistico/ -> 10-DASHIBOARDs/)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    fallback = os.path.normpath(os.path.join(script_dir, "..", ".."))
    return fallback

PASTA_DASHBOARDS = _encontrar_pasta()
ARQUIVO_BASE    = os.path.join(PASTA_DASHBOARDS, "arquivo_base.xls")
QUADRO_ENVIOS   = os.path.join(PASTA_DASHBOARDS, "Quadro de Envios.xlsx")
PASTA_BACKUP    = os.path.join(PASTA_DASHBOARDS, "backups")
ABA_LANCAMENTOS = "Lançamentos"

# CNPJ Remetente → Empresa
CNPJ_EMPRESA_MAP = {
    "08.458.084/0001-13": "IAB",
    "57.638.518/0002-53": "VERBUM",
    "57.638.518/0001-72": "VERBUM",
}


# Normalização de Natureza de Operação (mesmo mapeamento do cleaner.py)
NATUREZA_MAP = {
    "NOTA DE VENDA DE PRODUTO PRODUZIDO":          "VENDA",
    "VENDA ENTREGA FUTURA - SIMPLES FAT":          "VENDA",
    "VENDAS DE PRODUTOS - MI":                     "VENDA",
    "1010100 - VENDAS DE PRODUTOS - MI":           "VENDA",
    "VENDA":                                       "VENDA",
    "VENDA CONSIGNADA":                            "VENDA",
    "REMESSA VENDA ENTREGA FUTURA":                "VENDA",
    "REMESSA DE BONIFICAÇÃO - SAIDA":              "BONIFICAÇÃO",
    "BONIFICACAO":                                 "BONIFICAÇÃO",
    "BONIFICAÇÃO":                                 "BONIFICAÇÃO",
    "2040100 - BONIFICACAO":                       "BONIFICAÇÃO",
    "DEVOLUÇÃO DE VENDA - NF PRÓPRIA":             "DEVOLUÇÃO",
    "DEVOLUÇÃO DE VENDA - NF TERCEIROS":           "DEVOLUÇÃO",
    "DEVOLUÇÃO DE REMESSA DE BONIFICAÇÃO":         "DEVOLUÇÃO",
    "DEVOLUÇÃO DE REMESSA - ENTREGA FUTURA":       "DEVOLUÇÃO",
    "REMESSA EM CONSIGNAÇÃO - P/ VENDA":           "CONSIGNAÇÃO",
    "SIMPLES REMESSA (OUTRAS)":                    "SIMPLES REMESSA",
    "RETORNO SIMPLES REMESSA (OUTRAS) PROPRIO":    "SIMPLES REMESSA",
    "ESTORNO VDA - ENTREGA FUTURA SIMPLES FAT":    "SIMPLES REMESSA",
    "TRANSFERENCIA ENTRE FILIAIS - SAIDA LIVROS PROPRIOS": "TRANSFERÊNCIA",
    "TRANSFERENCIA ENTRE FILIAIS - SAIDA":         "TRANSFERÊNCIA",
    "AJUSTE DE ESTOQUE - SAIDA/DESCARTE":          "OUTRAS SAÍDAS",
    "AJUSTE DE ESTOQUE - SAÍDA DE ESTOQUE COM TERCEIROS": "OUTRAS SAÍDAS",
    "NOTA DE VENDA DE SERVIÇOS COALA":             "OUTRAS SAÍDAS",
}

# ── MAPEAMENTO UF → REGIÃO ──
UF_REGIAO = {
    "AC":"NORTE","AL":"NORDESTE","AM":"NORTE","AP":"NORTE","BA":"NORDESTE",
    "CE":"NORDESTE","DF":"CENTRO-OESTE","ES":"SUDESTE","EX":"EXTERIOR",
    "GO":"CENTRO-OESTE","MA":"NORDESTE","MG":"SUDESTE","MS":"CENTRO-OESTE",
    "MT":"CENTRO-OESTE","PA":"NORTE","PB":"NORDESTE","PE":"NORDESTE",
    "PI":"NORDESTE","PR":"SUL","RJ":"SUDESTE","RN":"NORDESTE","RO":"NORTE",
    "RR":"NORTE","RS":"SUL","SC":"SUL","SE":"NORDESTE","SP":"SUDESTE",
    "TO":"NORTE",
}

# ── CNPJS/CPFS CLASSIFICADOS ──
# Coordenadores (histórico do Quadro de Envios)
CNPJS_COORD = {
    "01031753427","04050264471","04640831412","05762129462",
    "07175228651","07725916792","08854762636","09477335746",
    "14891051000165","21728283353","41761871153","50649361091",
    "61226300634","76065227315","88861767087",
}

# Mapa CNPJ → Tipo de Cliente
# Construído a partir do histórico do Quadro de Envios + tabela de correspondência
# CPF coordenador identificado pelos dígitos
MAPA_CNPJ_TIPO = {"25140579000119": "MUNICÍPIO", "88185020000125": "MUNICÍPIO", "30775851000177": "MUNICÍPIO", "18602011000107": "MUNICÍPIO", "06554919000113": "MUNICÍPIO", "37464831000124": "MUNICÍPIO", "07598634000137": "MUNICÍPIO", "04204318000145": "MUNICÍPIO", "83021808000182": "MUNICÍPIO", "30017572000144": "MUNICÍPIO", "01612579000106": "MUNICÍPIO", "06138856000104": "MUNICÍPIO", "90152240000102": "MUNICÍPIO", "05943030000155": "MUNICÍPIO", "12151993000181": "MUNICÍPIO", "06554836000114": "MUNICÍPIO", "18602037000155": "MUNICÍPIO", "06554869000598": "MUNICÍPIO", "87613006000112": "MUNICÍPIO", "92406438000192": "MUNICÍPIO", "75462820000102": "MUNICÍPIO", "22646525000131": "MUNICÍPIO", "92406248000175": "MUNICÍPIO", "13098181000192": "MUNICÍPIO", "92000231000113": "MUNICÍPIO", "01612622000133": "MUNICÍPIO", "06554885000157": "MUNICÍPIO", "01612619000110": "MUNICÍPIO", "06728240000193": "MUNICÍPIO", "90221565000191": "MUNICÍPIO", "30734376000190": "MUNICÍPIO", "37465283000157": "MUNICÍPIO", "06084696000168": "MUNICÍPIO", "31159154000154": "MUNICÍPIO", "31907632000167": "MUNICÍPIO", "06554356000153": "MUNICÍPIO", "06554364000280": "MUNICÍPIO", "876612800000141": "MUNICÍPIO", "87590998000100": "MUNICÍPIO", "06072005000106": "MUNICÍPIO", "16516152000181": "MUNICÍPIO", "06554018000111": "MUNICÍPIO", "13128814001049": "MUNICÍPIO", "30567122000125": "MUNICÍPIO", "31015782000166": "MUNICÍPIO", "01612569000170": "MUNICÍPIO", "30601613000145": "MUNICÍPIO", "07569205000131": "MUNICÍPIO", "97229181000164": "MUNICÍPIO", "01613309000110": "MUNICÍPIO", "03431812000180": "MUNICÍPIO", "82930181000110": "MUNICÍPIO", "06554760000127": "MUNICÍPIO", "18137927000133": "MUNICÍPIO", "73357469000156": "MUNICÍPIO", "06101117000148": "MUNICÍPIO", "17749896000109": "MUNICÍPIO", "16726028000140": "MUNICÍPIO", "08917080000156": "MUNICÍPIO", "06554919000103": "MUNICÍPIO", "83102244000102": "MUNICÍPIO", "13098181000182": "MUNICÍPIO", "18671271000134": "MUNICÍPIO", "82926536000105": "MUNICÍPIO", "13104112000134": "MUNICÍPIO", "30969129000173": "MUNICÍPIO", "83102764000115": "MUNICÍPIO", "20622890000180": "MUNICÍPIO", "82909409000190": "MUNICÍPIO", "01006232000110": "MUNICÍPIO", "92000215000120": "MUNICÍPIO", "52281965000101": "MUNICÍPIO", "82916800000111": "MUNICÍPIO", "06553614000187": "MUNICÍPIO", "42774281000180": "MUNICÍPIO", "46384111000140": "MUNICÍPIO", "30258237000138": "MUNICÍPIO", "30934924000126": "MUNICÍPIO", "45739083000173": "MUNICÍPIO", "87612800000141": "MUNICÍPIO", "06554802000120": "MUNICÍPIO", "12200275000158": "MUNICÍPIO", "18584961000156": "MUNICÍPIO", "34693564000179": "MUNICÍPIO", "06554026000168": "MUNICÍPIO", "06553697000104": "MUNICÍPIO", "11408097000192": "MUNICÍPIO", "06896534000124": "MUNICÍPIO", "06553887000121": "MUNICÍPIO", "06554943000142": "MUNICÍPIO", "03074245000151": "MUNICÍPIO", "25173187000156": "MUNICÍPIO", "06554232000178": "MUNICÍPIO", "30371112000110": "MUNICÍPIO", "70946009000175": "MUNICÍPIO", "30676302000145": "MUNICÍPIO", "46137477000114": "MUNICÍPIO", "06200267000109": "MUNICÍPIO", "06156160000100": "MUNICÍPIO", "30044520000167": "MUNICÍPIO", "63042239000144": "PESSOA JURÍDICA", "58993577000121": "MUNICÍPIO", "3455397600103": "ESCOLA", "23819759000104": "ESCOLA", "55241543000192": "ESCOLA", "04455445000117": "ESCOLA", "23103586000115": "ESCOLA", "37288612000131": "ESCOLA", "52405166000191": "ESCOLA", "37544575000185": "ESCOLA", "63019772000438": "ESCOLA", "35936521000130": "ESCOLA", "51006232000198": "ESCOLA", "26462477000182": "ESCOLA", "14213556000170": "ESCOLA", "16582288000190": "ESCOLA", "03690751000175": "ESCOLA", "10548157000100": "ESCOLA", "51910155000104": "ESCOLA", "29761234000133": "ESCOLA", "24412991000188": "ESCOLA", "33466264000194": "ESCOLA", "47942708000125": "ESCOLA", "10548157000290": "ESCOLA", "16582288000351": "ESCOLA", "36195301000165": "ESCOLA", "00497704000111": "ESCOLA", "50596790001160": "ESCOLA", "220706438000144": "ESCOLA", "12947194000116": "ESCOLA", "54736220000107": "ESCOLA", "55952395000114": "ESCOLA", "43932923000194": "ESCOLA", "50703960000196": "ESCOLA", "18926708000133": "ESCOLA", "54392338000165": "ESCOLA", "30767703000100": "ESCOLA", "45359586000113": "ESCOLA", "27557042000184": "ESCOLA", "50964171000109": "ESCOLA", "20231133000186": "ESCOLA", "43905888000114": "ESCOLA", "20872280000135": "ESCOLA", "11369387000174": "ESCOLA", "08335714000162": "ESCOLA", "05209619000124": "ESCOLA", "48553564000188": "ESCOLA", "57370596000139": "ESCOLA", "43437967878": "ESCOLA", "52187570000136": "ESCOLA", "05906146000114": "ESCOLA", "47008618000161": "ESCOLA", "46008618000161": "ESCOLA", "07535734000114": "ESCOLA", "60613552000105": "ESCOLA", "27429498000169": "ESCOLA", "00580852000103": "ESCOLA", "11233342000178": "ESCOLA", "57516228000156": "ESCOLA", "02231133000186": "ESCOLA", "60849128000156": "ESCOLA", "14208959000121": "ESCOLA", "07112308000178": "ESCOLA", "33544370001110": "ESCOLA", "04071106000137": "ESCOLA", "60454342000103": "ESCOLA", "49870101000102": "ESCOLA", "56228516000142": "ESCOLA", "62211580000113": "ESCOLA", "55145110000133": "ESCOLA", "28398228000109": "ESCOLA", "52579682000132": "ESCOLA", "03760938000106": "ESCOLA", "61805204000194": "ESCOLA", "50303252000148": "ESCOLA", "55903254000101": "ESCOLA", "62719230000162": "ESCOLA", "27929089000202": "ESCOLA", "10548157000371": "ESCOLA", "56875117000173": "ESCOLA", "56237499000100": "ESCOLA", "12803829000101": "ESCOLA", "34553976000103": "ESCOLA", "45898691000201": "ESCOLA", "59968697000131": "ESCOLA", "10548157000452": "ESCOLA", "50303252000168": "ESCOLA", "50114307000191": "ESCOLA", "20736476000100": "ESCOLA", "56349104000160": "ESCOLA", "54348190000161": "ESCOLA", "10847762001315": "ESCOLA", "40301427000144": "ESCOLA", "51304748000119": "ESCOLA", "56049029000112": "ESCOLA", "14891051000165": "COORDENADOR/EVENTOS", "01031753427": "COORDENADOR/EVENTOS", "08854762636": "COORDENADOR/EVENTOS", "21728283353": "COORDENADOR/EVENTOS", "09477335746": "COORDENADOR/EVENTOS", "50649361091": "COORDENADOR/EVENTOS", "61226300634": "COORDENADOR/EVENTOS", "05762129462": "COORDENADOR/EVENTOS", "88861767087": "COORDENADOR/EVENTOS", "76065227315": "COORDENADOR/EVENTOS", "07725916792": "COORDENADOR/EVENTOS", "04640831412": "COORDENADOR/EVENTOS", "04050264471": "COORDENADOR/EVENTOS", "07175228651": "COORDENADOR/EVENTOS", "41761871153": "COORDENADOR/EVENTOS", "19004863000165": "E-BOOK", "42215329000110": "PESSOA JURÍDICA", "43252294000151": "PESSOA JURÍDICA", "06915161000191": "PESSOA JURÍDICA", "52997491000190": "PESSOA JURÍDICA", "18202870000109": "PESSOA JURÍDICA", "16561461000173": "PESSOA JURÍDICA", "09372101000168": "PESSOA JURÍDICA", "39643992000100": "PESSOA JURÍDICA", "46696975000105": "PESSOA JURÍDICA", "26821912000118": "PESSOA JURÍDICA", "08463170000386": "PESSOA JURÍDICA", "08463170000467": "PESSOA JURÍDICA", "08458084000385": "PESSOA JURÍDICA", "42416634000170": "PESSOA JURÍDICA", "29884158000153": "PESSOA JURÍDICA", "61700977000106": "PESSOA JURÍDICA", "35159835000173": "PESSOA JURÍDICA", "08458084000113": "PESSOA JURÍDICA", "57638518000172": "PESSOA JURÍDICA", "39791674000197": "PESSOA JURÍDICA", "40125545000149": "PESSOA JURÍDICA", "258843140683": "PESSOA JURÍDICA", "00943162000163": "PESSOA JURÍDICA", "01804403000156": "PESSOA JURÍDICA", "23468555000249": "PESSOA JURÍDICA", "31172205000188": "PESSOA JURÍDICA", "11768154000144": "PESSOA JURÍDICA", "55713639000106": "PESSOA JURÍDICA", "28239060000180": "PESSOA JURÍDICA", "00752439000170": "PESSOA JURÍDICA", "45027375000183": "PESSOA JURÍDICA", "06967669000133": "PESSOA JURÍDICA", "60701521000106": "PESSOA JURÍDICA", "57488056000154": "PESSOA JURÍDICA", "05774744000187": "PESSOA JURÍDICA", "57638518000253": "PESSOA JURÍDICA", "61876263000153": "PESSOA JURÍDICA", "50857561000180": "PESSOA JURÍDICA", "78796778000146": "PESSOA JURÍDICA", "05031043000158": "PESSOA JURÍDICA", "06878967000157": "PESSOA JURÍDICA", "21295545000142": "PESSOA JURÍDICA", "00235200000123": "PESSOA JURÍDICA", "61265278000185": "PESSOA JURÍDICA", "21347139000186": "PESSOA JURÍDICA", "61062202000152": "PESSOA JURÍDICA", "00702446000168": "PESSOA JURÍDICA", "14101600000150": "PESSOA JURÍDICA", "37547218000170": "PESSOA JURÍDICA", "07707869000110": "PESSOA JURÍDICA", "01994969000198": "PESSOA JURÍDICA", "43586967000100": "PESSOA JURÍDICA", "02716515395": "PESSOA JURÍDICA", "08458084000547": "PESSOA JURÍDICA", "58479262001302": "PESSOA JURÍDICA", "01613315000177": "PESSOA JURÍDICA", "80618051000200": "PESSOA JURÍDICA", "29980143000199": "PESSOA JURÍDICA", "06200544000183": "PESSOA JURÍDICA", "05554018000159": "PESSOA JURÍDICA", "17803663000147": "PESSOA JURÍDICA", "01434589000956": "PESSOA JURÍDICA", "17881485000849": "PESSOA JURÍDICA", "78636974000315": "PESSOA JURÍDICA", "86552809000222": "PESSOA JURÍDICA", "04795928000837": "PESSOA JURÍDICA", "33667098000194": "PESSOA JURÍDICA", "53282760000286": "PESSOA JURÍDICA", "17881485000172": "PESSOA JURÍDICA", "31504244000135": "PESSOA JURÍDICA", "52576998000170": "PESSOA JURÍDICA", "13138832000110": "PESSOA JURÍDICA", "06845408000302": "PESSOA JURÍDICA", "92023159003408": "PESSOA JURÍDICA", "33618984000209": "PESSOA JURÍDICA", "62959591000186": "PESSOA JURÍDICA", "84467562000130": "PESSOA JURÍDICA", "63806400000109": "PESSOA JURÍDICA", "92880962000281": "PESSOA JURÍDICA", "51722148000170": "PESSOA JURÍDICA", "29292673000144": "PESSOA JURÍDICA", "01655397000112": "PESSOA JURÍDICA", "12678115000119": "PESSOA JURÍDICA", "53416921000520": "PESSOA JURÍDICA", "07526442000115": "PESSOA JURÍDICA", "35829994000219": "PESSOA JURÍDICA", "06296071000160": "PESSOA JURÍDICA", "54074205000140": "PESSOA JURÍDICA", "24207268000167": "PESSOA JURÍDICA", "54131610000153": "PESSOA JURÍDICA", "58191008000162": "PESSOA JURÍDICA", "01305652000105": "PESSOA JURÍDICA", "41406851000116": "PESSOA JURÍDICA", "55399805000141": "PESSOA JURÍDICA", "11952647000130": "PESSOA JURÍDICA", "44859254000135": "PESSOA JURÍDICA", "74058017000136": "PESSOA JURÍDICA", "13074007000108": "PESSOA JURÍDICA", "07472640000143": "PESSOA JURÍDICA", "05479860000173": "PESSOA JURÍDICA", "30717813000753": "PESSOA JURÍDICA"}

def _carregar_cnpjs_classificados():
    """
    Retorna mapa atualizado lendo o Quadro e a tabela de escolas em tempo real.
    Combina com MAPA_CNPJ_TIPO (base histórica) — dados em tempo real têm prioridade.
    """
    import re as _re
    mapa = dict(MAPA_CNPJ_TIPO)  # cópia da base histórica
    try:
        import pandas as _pd
        # Ler Quadro de Envios — atualizar com registros mais recentes
        qe = _pd.read_excel(QUADRO_ENVIOS, sheet_name="Lançamentos",
                            engine="openpyxl", dtype=str,
                            usecols=["CNPJ / CPF","Tipo de Cliente"])
        for tipo in ["MUNICÍPIO","ESCOLA","COORDENADOR/EVENTOS","E-BOOK","PESSOA JURÍDICA"]:
            subset = qe[qe["Tipo de Cliente"].str.strip().str.upper()==tipo]["CNPJ / CPF"].dropna()
            for cnpj in subset:
                dig = _re.sub(r"\D","",str(cnpj))
                if dig and len(dig) > 5:
                    mapa[dig] = tipo
        # Tabela de escolas tem prioridade máxima
        pasta = os.path.dirname(QUADRO_ENVIOS)
        arq_tc = os.path.join(pasta, "tabela_correspondencia_nome_clientes.xlsx")
        if os.path.exists(arq_tc):
            tc = _pd.read_excel(arq_tc, engine="openpyxl", skiprows=2, header=0)
            tc.columns = (["cnpj","nome_atual","nome_pad","municipio","uf","regiao","obs","ativo"] +
                          [f"x{i}" for i in range(max(0, len(tc.columns)-8))])
            for cnpj in tc["cnpj"].dropna():
                dig = _re.sub(r"\D","",str(cnpj))
                if dig and len(dig) > 5:
                    mapa[dig] = "ESCOLA"
    except Exception as e:
        log.warning(f"Aviso ao carregar CNPJs: {e}")
    return mapa


# ── FUNÇÕES DE CLASSIFICAÇÃO ──
def classificar_tipo_cliente(cnpj_cpf, mapa_tipos):
    """
    Classifica o tipo de cliente pelo CNPJ/CPF.
    Retorna (tipo, cinza) onde cinza=True significa célula cinza.
    """
    import re as _re
    if not cnpj_cpf: return "", True
    digitos = _re.sub(r"\D","", str(cnpj_cpf))
    if len(digitos) == 11:             # CPF — Pessoa Física ou Coordenador
        if digitos in CNPJS_COORD or mapa_tipos.get(digitos) == "COORDENADOR/EVENTOS":
            return "COORDENADOR/EVENTOS", False
        return "PESSOA FÍSICA", False
    elif len(digitos) == 14:           # CNPJ
        tipo = mapa_tipos.get(digitos)
        if tipo:
            return tipo, False
        return "PESSOA JURÍDICA", False   # CNPJ não mapeado → PJ
    return "", True                       # Desconhecido → branco + cinza


def calcular_ano_ref(data_coleta):
    """01/10/ANO → ANO+1 | 01/01 a 30/09/ANO → ANO"""
    try:
        if not data_coleta or str(data_coleta) == "NaT": return ""
        d = pd.Timestamp(data_coleta)
        return str(d.year + 1) if d.month >= 10 else str(d.year)
    except: return ""


def calcular_regiao(uf):
    """Retorna a região baseada na UF."""
    if not uf: return ""
    return UF_REGIAO.get(str(uf).strip().upper(), "")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger()


def resolver_empresa(cnpj_remetente):
    cnpj = str(cnpj_remetente).strip() if cnpj_remetente else ""
    return CNPJ_EMPRESA_MAP.get(cnpj, "")


def montar_chave(data, nf, cnpj_dest, empresa=""):
    """Chave: AAAA-MM-DD|NF|CNPJ_digitos|EMPRESA"""
    try:
        data_str = pd.Timestamp(data).strftime("%Y-%m-%d") if not pd.isna(data) else ""
    except:
        data_str = ""
    try:
        nf_str = str(int(float(nf))) if not pd.isna(nf) else ""
    except:
        nf_str = ""
    cnpj_str = "".join(c for c in str(cnpj_dest) if c.isdigit()) if cnpj_dest else ""
    emp_str  = str(empresa).strip().upper()
    return f"{data_str}|{nf_str}|{cnpj_str}|{emp_str}"


def fazer_backup():
    os.makedirs(PASTA_BACKUP, exist_ok=True)
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = os.path.join(PASTA_BACKUP, f"Quadro_de_Envios_backup_{ts}.xlsx")
    shutil.copy2(QUADRO_ENVIOS, destino)
    log.info(f"Backup salvo: {destino}")
    return destino


def ler_arquivo_base():
    log.info(f"Lendo: {ARQUIVO_BASE}")
    df = pd.read_excel(ARQUIVO_BASE, engine="xlrd", skiprows=2, header=0, dtype=str)
    df.dropna(how="all", inplace=True)
    df["DT_EMISSAO_NF"]      = pd.to_datetime(df["DT_EMISSAO_NF"],      errors="coerce")
    df["NRO_NFE"]            = pd.to_numeric(df["NRO_NFE"],             errors="coerce")
    df["VALOR_TOTAL_NOTA"]   = pd.to_numeric(df["VALOR_TOTAL_NOTA"],    errors="coerce")
    df["VALOR_BASE_CALCULO"] = pd.to_numeric(df["VALOR_BASE_CALCULO"],  errors="coerce")
    df["VALOR_ICMS"]         = pd.to_numeric(df["VALOR_ICMS"],          errors="coerce")
    df["VALOR_DIFAL"]        = pd.to_numeric(df["VALOR_DIFAL"],         errors="coerce")
    df["VALOR_FCP"]          = pd.to_numeric(df["VALOR_FCP"],           errors="coerce")
    df["PESO"]               = pd.to_numeric(df["PESO"],                errors="coerce")
    df["QTD_VOLUME"]         = pd.to_numeric(df["QTD_VOLUME"],          errors="coerce")
    log.info(f"Arquivo base: {len(df)} registros")
    return df


def ler_quadro_envios():
    log.info(f"Lendo: {QUADRO_ENVIOS}")
    df = pd.read_excel(QUADRO_ENVIOS, sheet_name=ABA_LANCAMENTOS,
                       engine="openpyxl", dtype=str)
    df["Data Coleta"] = pd.to_datetime(df["Data Coleta"], errors="coerce")
    df["NF"]          = pd.to_numeric(df["NF"], errors="coerce")
    log.info(f"Quadro de Envios: {len(df)} linhas")
    return df


def filtrar_novas(df_base, df_quadro):
    # Montar chaves do Quadro (Data + NF + CNPJ + Empresa)
    # Gera chave COM e SEM empresa para cobrir casos onde Empresa está vazia no Quadro
    chaves_quadro = set()
    for _, row in df_quadro.iterrows():
        cnpj = row.get("CNPJ / CPF")
        data = row.get("Data Coleta")
        nf   = row.get("NF")
        emp  = str(row.get("Empresa", "")).strip().upper()
        # Chave com empresa
        chaves_quadro.add(montar_chave(data, nf, cnpj, emp))
        # Chave sem empresa (para registros antigos sem empresa preenchida)
        chaves_quadro.add(montar_chave(data, nf, cnpj, ""))

    log.info(f"Chaves únicas no Quadro: {len(chaves_quadro)}")

    # Identificar novas
    novas = []
    for _, row in df_base.iterrows():
        emp = resolver_empresa(row.get("CNPJ_REMETENTE", ""))
        ch  = montar_chave(row.get("DT_EMISSAO_NF"), row.get("NRO_NFE"),
                           row.get("CPF_CNPJ_DESTINATARIO"), emp)
        if ch not in chaves_quadro:
            novas.append(row)

    log.info(f"Registros novos a inserir: {len(novas)}")
    return novas


def mapear_linha(row_base, tipo_cli="", cinza_tc=False):
    data_coleta = row_base.get("DT_EMISSAO_NF")
    try:
        data_coleta = pd.Timestamp(data_coleta) if not pd.isna(data_coleta) else None
    except:
        data_coleta = None

    nf_val = row_base.get("NRO_NFE")
    try:
        nf_val = int(float(nf_val)) if not pd.isna(nf_val) else None
    except:
        nf_val = None

    # TIPO_OPERACAO do arquivo_base → NATUREZA DA OPERAÇÃO do Quadro (nome original)
    # A normalização acontece no ETL (cleaner.py) ao gerar o dashboard
    natureza_raw = str(row_base.get("TIPO_OPERACAO", "")).strip()
    natureza = natureza_raw if natureza_raw and natureza_raw != "nan" else ""

    empresa = resolver_empresa(row_base.get("CNPJ_REMETENTE", ""))
    cnpj_cpf = str(row_base.get("CPF_CNPJ_DESTINATARIO", "")).strip()
    uf = str(row_base.get("UF_DESTINO", "")).strip().upper()
    # tipo_cli e cinza_tc vêm como parâmetros da função (calculados antes da chamada)
    ano_ref  = calcular_ano_ref(data_coleta)
    regiao   = calcular_regiao(uf)

    # Ano e Mês da data de coleta
    try:
        ano_col = int(data_coleta.year)  if data_coleta else None
        mes_col = int(data_coleta.month) if data_coleta else None
    except: ano_col = mes_col = None

    return {
        "Data Coleta":          data_coleta,
        "Ano":                  ano_col,
        "Mês":                  mes_col,
        "CNPJ / CPF":           cnpj_cpf,
        "Cliente":              str(row_base.get("CLIENTE_DESTINATARIO", "")).strip(),
        "NF":                   nf_val,
        "Empresa":              empresa,
        "NATUREZA DA OPERAÇÃO": natureza,
        "Vr NF":                abs(float(row_base.get("VALOR_TOTAL_NOTA") or 0)),
        "Base Cálculo ICM":     abs(float(row_base.get("VALOR_BASE_CALCULO") or 0)),
        "ICMS":                 abs(float(row_base.get("VALOR_ICMS") or 0)),
        "Difal":                abs(float(row_base.get("VALOR_DIFAL") or 0)),
        "FCP":                 abs(float(row_base.get("VALOR_FCP") or 0)),
        "Transportadora":       (lambda t: "PRÓPRIO" if t.upper() in ("NAO INFORMADO","NÃO INFORMADO","") else t)(str(row_base.get("TRANSPORTADORA","")).strip()),
        "Peso (kg)":            row_base.get("PESO"),
        "Volumes":              row_base.get("QTD_VOLUME"),
        "Municipio":            str(row_base.get("CIDADE_DESTINO", "")).strip(),
        "UF":                   uf,
        "Região":               regiao,
        "Tipo de Cliente":      tipo_cli,
        "Ano Ref":              ano_ref,
        "Cons Analise":         "Sim",
        "País":                 "BRASIL",
        # Marcadores para células cinza
        "_cinza_tipo_cli":      cinza_tc,
        "_cinza_cfop":          True,      # Cfop sempre cinza
    }


def inserir_no_quadro(linhas_novas):
    if not linhas_novas:
        log.info("Nenhuma linha nova para inserir.")
        return 0

    log.info(f"Abrindo {QUADRO_ENVIOS} para inserção...")
    wb = openpyxl.load_workbook(QUADRO_ENVIOS)
    ws = wb[ABA_LANCAMENTOS]

    # Mapa coluna → índice
    header = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header[str(cell.value).strip()] = col_idx

    # Última linha com dados
    ultima_linha = ws.max_row
    while ultima_linha > 1:
        if any(ws.cell(row=ultima_linha, column=c).value
               for c in range(1, len(header)+2)):
            break
        ultima_linha -= 1

    proxima = ultima_linha + 1
    log.info(f"Inserindo a partir da linha {proxima}")

    from openpyxl.styles import PatternFill
    FILL_CINZA = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Carregar mapa CNPJ → Tipo de Cliente (histórico + tempo real)
    mapa_tipos = _carregar_cnpjs_classificados()
    log.info(f"Mapa de tipos carregado: {len(mapa_tipos)} CNPJs classificados")

    inseridos = 0
    ignorados = 0
    for row_base in linhas_novas:
        # Ignorar linhas inválidas do arquivo_base
        nf_val   = row_base.get("NRO_NFE", "")
        cnpj_val = row_base.get("CPF_CNPJ_DESTINATARIO", "")
        cli_val  = row_base.get("CLIENTE_DESTINATARIO", "")
        import math as _math
        def _invalido(v):
            if v is None: return True
            s = str(v).strip().lower()
            if s in ("", "nan", "none", "nat"): return True
            try:
                f = float(v)
                return _math.isnan(f)
            except: return False

        if _invalido(nf_val) or _invalido(cnpj_val) or _invalido(cli_val):
            ignorados += 1
            continue

        # Ignorar NF com valor 0 ou negativo
        try:
            nf_num = int(float(nf_val))
            if nf_num <= 0:
                ignorados += 1
                continue
        except:
            ignorados += 1
            continue

        # Classificar tipo de cliente antes de mapear
        cnpj = str(cnpj_val).strip()
        tipo_cli, cinza_tc = classificar_tipo_cliente(cnpj, mapa_tipos)

        mapeado = mapear_linha(row_base, tipo_cli=tipo_cli, cinza_tc=cinza_tc)

        # Campos marcadores (não são colunas reais)
        cinza_tipo = mapeado.pop("_cinza_tipo_cli", False)
        cinza_cfop = mapeado.pop("_cinza_cfop", False)

        for nome_col, valor in mapeado.items():
            if nome_col not in header: continue
            cell = ws.cell(row=proxima, column=header[nome_col])
            if isinstance(valor, pd.Timestamp):
                cell.value = valor.to_pydatetime()
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = valor

        # Pintar células cinza
        if cinza_cfop and "Cfop" in header:
            ws.cell(row=proxima, column=header["Cfop"]).fill = FILL_CINZA
        if cinza_tipo and "Tipo de Cliente" in header:
            ws.cell(row=proxima, column=header["Tipo de Cliente"]).fill = FILL_CINZA

        proxima  += 1
        inseridos += 1

    wb.save(QUADRO_ENVIOS)
    if ignorados:
        log.info(f"{ignorados} linhas ignoradas (sem NF, CNPJ ou Cliente)")
    log.info(f"{inseridos} linhas inseridas e salvas")
    return inseridos


def main():
    inicio = datetime.now()
    print()
    print("=" * 55)
    print("  IMPORTACAO ARQUIVO BASE -> QUADRO DE ENVIOS")
    print(f"  {inicio.strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 55)
    print()

    for arq in [ARQUIVO_BASE, QUADRO_ENVIOS]:
        if not os.path.exists(arq):
            log.error(f"Arquivo nao encontrado: {arq}")
            sys.exit(1)

    fazer_backup()
    df_base   = ler_arquivo_base()
    df_quadro = ler_quadro_envios()
    novas     = filtrar_novas(df_base, df_quadro)
    total     = inserir_no_quadro(novas)

    duracao = (datetime.now() - inicio).total_seconds()
    print()
    print("=" * 55)
    if total > 0:
        print(f"  OK  {total} registros novos inseridos no Quadro")
    else:
        print(f"  OK  Nenhum registro novo - Quadro ja atualizado")
    print(f"  Tempo: {duracao:.1f}s")
    print("=" * 55)
    print()


if __name__ == "__main__":
    main()
